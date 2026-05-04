"""Export extractions to Excel.

Two sheets: ``Inventory`` (one row per contract, flat KPI columns) and
``Evidence`` (one row per extracted quote with page + file). Rows where
``review_required`` is True are highlighted so reviewers can triage.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema import ContractExtraction


INVENTORY_HEADERS = [
    "contract_id",
    "review_required",
    "confidence",
    "source_files",
    "base_package_summary",
    "base_package_quote",
    "base_package_page",
    "tm_addon_count",
    "tm_addons",
    "risk_count",
    "risk_severities",
    "risk_types",
    "patch_required",
    "patch_pricing_model",
    "patch_scope",
    "patch_end_date",
    "patch_page",
    "not_found",
    "validation_warnings",
    "reviewer_notes",
]

EVIDENCE_HEADERS = [
    "contract_id",
    "field",
    "quote",
    "source_file",
    "source_page",
    "source_section",
    "extra",
]


def write_excel(extractions: Iterable[ContractExtraction], path: Path) -> None:
    wb = Workbook()
    inv = wb.active
    inv.title = "Inventory"
    inv.append(INVENTORY_HEADERS)

    ev = wb.create_sheet("Evidence")
    ev.append(EVIDENCE_HEADERS)

    for extraction in extractions:
        inv.append(_inventory_row(extraction))
        for row in _evidence_rows(extraction):
            ev.append(row)

    _style_header(inv)
    _style_header(ev)
    _autosize(inv)
    _autosize(ev)
    _highlight_review_rows(inv)
    _highlight_low_confidence(inv)

    wb.save(path)


def _inventory_row(e: ContractExtraction) -> list:
    addons = "; ".join(a.service for a in e.tm_addons)
    severities = ", ".join(sorted({f.severity.value for f in e.commercial_risk_flags}))
    risk_types = ", ".join(sorted({f.risk_type.value for f in e.commercial_risk_flags}))
    return [
        e.contract_id,
        "yes" if e.review_required else "no",
        e.confidence.value,
        "; ".join(Path(p).name for p in e.source_files),
        e.base_package.scope_summary or "",
        e.base_package.definition_quote or "",
        e.base_package.source_page or "",
        len(e.tm_addons),
        addons,
        len(e.commercial_risk_flags),
        severities,
        risk_types,
        _bool_str(e.patch_management.required),
        e.patch_management.pricing_model.value,
        e.patch_management.scope_summary or "",
        e.patch_management.end_date or "",
        e.patch_management.source_page or "",
        "; ".join(e.not_found),
        " | ".join(e.validation_warnings),
        e.reviewer_notes or "",
    ]


def _evidence_rows(e: ContractExtraction) -> list[list]:
    rows: list[list] = []
    if e.base_package.definition_quote:
        rows.append([
            e.contract_id,
            "base_package",
            e.base_package.definition_quote,
            e.base_package.source_file or "",
            e.base_package.source_page or "",
            e.base_package.source_section or "",
            "",
        ])
    for addon in e.tm_addons:
        rows.append([
            e.contract_id,
            f"tm_addon: {addon.service}",
            addon.quote,
            addon.source_file or "",
            addon.source_page or "",
            addon.source_section or "",
            addon.rate_card_ref or "",
        ])
    for flag in e.commercial_risk_flags:
        rows.append([
            e.contract_id,
            f"risk: {flag.risk_type.value} ({flag.severity.value})",
            flag.evidence_quote,
            flag.source_file or "",
            flag.source_page or "",
            flag.source_section or "",
            flag.rationale,
        ])
    if e.patch_management.evidence_quote:
        rows.append([
            e.contract_id,
            "patch_management",
            e.patch_management.evidence_quote,
            e.patch_management.source_file or "",
            e.patch_management.source_page or "",
            e.patch_management.source_section or "",
            f"pricing={e.patch_management.pricing_model.value}; end={e.patch_management.end_date or ''}",
        ])
    return rows


def _bool_str(value: bool | None) -> str:
    if value is None:
        return ""
    return "yes" if value else "no"


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="DDDDDD")
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = bold


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, column in enumerate(ws.columns, start=1):
        width = 12
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _highlight_review_rows(ws) -> None:
    review_fill = PatternFill("solid", fgColor="F4B084")  # orange — needs review
    review_col = INVENTORY_HEADERS.index("review_required") + 1
    for row in ws.iter_rows(min_row=2):
        if row[review_col - 1].value == "yes":
            for c in row:
                c.fill = review_fill


def _highlight_low_confidence(ws) -> None:
    yellow = PatternFill("solid", fgColor="FFF2CC")
    red = PatternFill("solid", fgColor="F8CBAD")
    confidence_col = INVENTORY_HEADERS.index("confidence") + 1
    review_col = INVENTORY_HEADERS.index("review_required") + 1
    for row in ws.iter_rows(min_row=2):
        # Don't override the orange review highlight
        if row[review_col - 1].value == "yes":
            continue
        cell = row[confidence_col - 1]
        if cell.value == "low":
            for c in row:
                c.fill = red
        elif cell.value == "medium":
            for c in row:
                c.fill = yellow
